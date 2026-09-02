"""
Cross-tenant isolation.

Phase 1 of the multi-client work. The RFP targets roughly 100 clients in 2026 scaling
to the wider Commercial Risk book in 2027 (Scale and Rollout Assumptions, p. 11), and
cross-client disclosure of Confidential claims data is the worst failure this system
has. These tests exist so that failure is caught by CI rather than by a client.

Two predicates are asserted throughout: the tenant (client_id) and the organisational
node scope (BR-001). The node scope tests already live in test_scope_negative.py; this
file is only about the tenant boundary, and deliberately holds the *role* constant
while varying only the client - so a pass cannot be explained by a privilege
difference.

Traceability: BR-001 (p. 38), F-CC-07, Scale and Rollout Assumptions (p. 11),
Application Profile data classification "Confidential" (p. 42).
"""
import io
import zipfile

import pytest
from conftest import (
    CLIENT_A, CLIENT_B, P1_CSUITE, P3_JFK_DIRECTOR, P5_BISTRO_MGR,
    T2_CSUITE, T2_REGIONAL, T2_STORE_MGR, auth,
)


# ── helpers ──────────────────────────────────────────────────────────────────

def claim_ids(client, token, **params):
    p = {"page_size": 100, **params}
    r = client.get("/api/v1/claims", params=p, headers=auth(token))
    assert r.status_code == 200, r.text
    return {i["aon_claim_id"] for i in r.json()["items"]}


def any_claim_of(client, token):
    ids = claim_ids(client, token)
    assert ids, "expected the tenant to have at least one claim"
    return sorted(ids)[0]


# ── the fixture itself ───────────────────────────────────────────────────────

def test_two_tenants_are_seeded(client, tokens):
    """Without a genuine second tenant every other test here is vacuous."""
    from app.db import query

    rows = query("SELECT client_id, COUNT(*) AS n FROM claims GROUP BY client_id")
    by_client = {r["client_id"]: r["n"] for r in rows}
    assert by_client.get(CLIENT_A, 0) > 0
    assert by_client.get(CLIENT_B, 0) > 0


def test_no_tenant_row_is_left_unassigned(client, tokens):
    """
    A NULL client_id is invisible to the two-predicate filter.

    That would present as missing data rather than as a fault, so it is asserted
    directly instead of being discovered through a confusing empty screen.
    """
    from app.db import query

    for table in ("org_nodes", "claims", "policies", "documents", "personas"):
        if table == "personas":
            # Persona 7 has no organisational node by design, so no tenant.
            n = query(
                "SELECT COUNT(*) AS n FROM personas "
                "WHERE client_id IS NULL AND org_node IS NOT NULL"
            )[0]["n"]
        else:
            n = query(f"SELECT COUNT(*) AS n FROM {table} WHERE client_id IS NULL")[0]["n"]
        assert n == 0, f"{table} has {n} rows with no tenant"


def test_token_carries_the_tenant(client, tokens):
    import jwt

    from app import settings

    for pid, expected in ((P1_CSUITE, CLIENT_A), (T2_CSUITE, CLIENT_B)):
        claims = jwt.decode(
            tokens[pid], settings.MOCK_JWT_SECRET, algorithms=["HS256"],
            audience=settings.OKTA_AUDIENCE,
        )
        assert claims["client_id"] == expected


# ── claims ───────────────────────────────────────────────────────────────────

def test_tenants_see_disjoint_claim_sets(client, tokens):
    a = claim_ids(client, tokens[P1_CSUITE])
    b = claim_ids(client, tokens[T2_CSUITE])
    assert a and b
    assert not (a & b), f"{len(a & b)} claims visible to both tenants"


def test_claim_detail_across_tenants_is_403(client, tokens):
    """Both directions, so a pass cannot come from one tenant simply having less."""
    a_claim = any_claim_of(client, tokens[P1_CSUITE])
    b_claim = any_claim_of(client, tokens[T2_CSUITE])

    assert client.get(f"/api/v1/claims/{a_claim}",
                      headers=auth(tokens[T2_CSUITE])).status_code == 403
    assert client.get(f"/api/v1/claims/{b_claim}",
                      headers=auth(tokens[P1_CSUITE])).status_code == 403


def test_cross_tenant_denial_is_audited(client, tokens):
    from app.db import query

    b_claim = any_claim_of(client, tokens[T2_CSUITE])
    client.get(f"/api/v1/claims/{b_claim}", headers=auth(tokens[P1_CSUITE]))
    rows = query(
        "SELECT * FROM audit_log WHERE outcome = 'denied' AND resource_id = :r",
        {"r": b_claim},
    )
    assert rows, "a cross-tenant attempt must leave an audit record"


def test_summary_totals_do_not_include_the_other_tenant(client, tokens):
    a = client.get("/api/v1/summary", headers=auth(tokens[P1_CSUITE])).json()
    b = client.get("/api/v1/summary", headers=auth(tokens[T2_CSUITE])).json()
    assert a["org_node"] == CLIENT_A
    assert b["org_node"] == CLIENT_B

    a_ids = claim_ids(client, tokens[P1_CSUITE])
    b_ids = claim_ids(client, tokens[T2_CSUITE])
    assert a["claim_count"] == len(a_ids)
    assert b["claim_count"] == len(b_ids)
    # Neither total may equal the combined set.
    assert a["claim_count"] != len(a_ids | b_ids)


def test_hierarchy_shows_only_the_callers_tenant(client, tokens):
    a = client.get("/api/v1/hierarchy", headers=auth(tokens[P1_CSUITE])).json()["nodes"]
    b = client.get("/api/v1/hierarchy", headers=auth(tokens[T2_CSUITE])).json()["nodes"]
    a_nodes = {n["org_node"] for n in a}
    b_nodes = {n["org_node"] for n in b}
    assert not (a_nodes & b_nodes)
    assert all(n.startswith(("CORP-HOSP", "LOC-JFK", "LOC-LHR", "LOC-SIN", "SITE-JFK",
                             "SITE-LHR", "SITE-SIN")) for n in a_nodes)
    assert all("NW" in n or n == CLIENT_B for n in b_nodes)


# ── a client-supplied tenant must never be honoured ─────────────────────────

@pytest.mark.parametrize("param", ["client_id", "tenant", "tenant_client_id", "org_node"])
def test_tenant_cannot_be_overridden_by_query_param(client, tokens, param):
    baseline = claim_ids(client, tokens[T2_STORE_MGR])
    escalated = claim_ids(client, tokens[T2_STORE_MGR], **{param: CLIENT_A})
    assert escalated == baseline


def test_tenant_cannot_be_overridden_by_header(client, tokens):
    baseline = client.get("/api/v1/claims", params={"page_size": 100},
                          headers=auth(tokens[T2_STORE_MGR])).json()["total"]
    headers = {**auth(tokens[T2_STORE_MGR]), "X-Client-Id": CLIENT_A}
    assert client.get("/api/v1/claims", params={"page_size": 100},
                      headers=headers).json()["total"] == baseline


def test_forged_tenant_claim_in_token_is_refused(client):
    """
    The hierarchy is authoritative for a node's tenant.

    A token asserting a different client than the node belongs to is refused rather
    than resolved in favour of either side - a mismatch means the two have diverged,
    and continuing would be guessing.
    """
    from app.auth.tokens import issue_mock_token

    forged = issue_mock_token({
        "persona_id": 99, "name": "Forged", "org_node": "SITE-NW-LEEDS",
        "client_id": CLIENT_A,          # node belongs to CLIENT_B
        "locale": "en-US", "groups_csv": "claims_viewer,claims_export",
    })
    r = client.get("/api/v1/claims", headers=auth(forged))
    assert r.status_code == 403


def test_node_without_a_tenant_is_refused(client):
    """An unassigned node cannot be scoped safely, so it is refused, not defaulted."""
    from app.auth.tokens import issue_mock_token
    from app.db import execute

    execute(
        "INSERT INTO org_nodes (org_node, parent_node, path, level, display_name, "
        "country_code, client_id) VALUES ('ORPHAN', NULL, '/ORPHAN/', 'corporate', "
        "'Orphan', 'US', NULL)"
    )
    tok = issue_mock_token({
        "persona_id": 98, "name": "Orphan", "org_node": "ORPHAN",
        "client_id": None, "locale": "en-US", "groups_csv": "claims_viewer",
    })
    assert client.get("/api/v1/claims", headers=auth(tok)).status_code == 403
    execute("DELETE FROM org_nodes WHERE org_node = 'ORPHAN'")


# ── documents (Pillar 1 still applies, plus the tenant boundary) ────────────

def test_documents_of_another_tenant_are_403(client, tokens):
    b_claim = any_claim_of(client, tokens[T2_CSUITE])
    r = client.get(f"/api/v1/claims/{b_claim}/documents", headers=auth(tokens[P1_CSUITE]))
    assert r.status_code == 403


def test_document_content_across_tenants_is_403(client, tokens):
    from app.db import query

    b_claim = any_claim_of(client, tokens[T2_CSUITE])
    docs = query(
        "SELECT doc_id FROM documents WHERE claim_id = :c AND audience = 'client_visible'",
        {"c": b_claim},
    )
    assert docs
    r = client.get(f"/api/v1/documents/{docs[0]['doc_id']}/content",
                   headers=auth(tokens[P1_CSUITE]))
    assert r.status_code == 403


# ── saved views ──────────────────────────────────────────────────────────────

def test_shared_view_does_not_cross_the_tenant_boundary(client, tokens):
    """
    Shared visibility travels upward within a tenant, and never sideways across one.

    The share rule reuses BR-001 rather than inventing a second model: a view is
    visible when its org_node sits inside the viewer's authorised scope. So a regional
    manager's view reaches the C-Suite above them, and the same view is invisible to
    the other tenant's C-Suite - who has an identical role and strictly more privilege.
    That second assertion is the one that matters here.
    """
    r = client.post("/api/v1/views",
                    json={"name": "Tenant B shared", "filters": {"status": "Open"},
                          "is_shared": True},
                    headers=auth(tokens[T2_REGIONAL]))
    assert r.status_code == 201
    vid = r.json()["view_id"]

    seen_within = {v["view_id"] for v in
                   client.get("/api/v1/views", headers=auth(tokens[T2_CSUITE])).json()["items"]}
    assert vid in seen_within, "a shared view should reach the scope above it in its own tenant"

    seen_across = {v["view_id"] for v in
                   client.get("/api/v1/views", headers=auth(tokens[P1_CSUITE])).json()["items"]}
    assert vid not in seen_across, "a shared view must not cross tenants"

    client.delete(f"/api/v1/views/{vid}", headers=auth(tokens[T2_REGIONAL]))


def test_view_created_in_one_tenant_is_stamped_with_it(client, tokens):
    from app.db import query_one

    r = client.post("/api/v1/views",
                    json={"name": "Stamp check", "filters": {"status": "Open"}},
                    headers=auth(tokens[T2_STORE_MGR]))
    vid = r.json()["view_id"]
    row = query_one("SELECT client_id FROM saved_views WHERE view_id = :v", {"v": vid})
    assert row["client_id"] == CLIENT_B
    client.delete(f"/api/v1/views/{vid}", headers=auth(tokens[T2_STORE_MGR]))


def test_cannot_delete_another_tenants_view(client, tokens):
    r = client.post("/api/v1/views",
                    json={"name": "B private", "filters": {"status": "Open"}},
                    headers=auth(tokens[T2_CSUITE]))
    vid = r.json()["view_id"]
    assert client.delete(f"/api/v1/views/{vid}",
                         headers=auth(tokens[P1_CSUITE])).status_code in (403, 404)
    client.delete(f"/api/v1/views/{vid}", headers=auth(tokens[T2_CSUITE]))


# ── filter options must not leak the other tenant's values ─────────────────

def test_filter_options_do_not_leak_across_tenants(client, tokens):
    a = client.get("/api/v1/claims-filter-options", headers=auth(tokens[P1_CSUITE])).json()
    b = client.get("/api/v1/claims-filter-options", headers=auth(tokens[T2_CSUITE])).json()
    # Reserve bounds are derived from the caller's own claims, so identical maxima
    # across two independently seeded tenants would indicate a shared query.
    assert a["reserve_max"] != b["reserve_max"]


# ── export carries the boundary into the generated file ────────────────────

def test_export_contains_only_the_callers_tenant(client, tokens):
    from openpyxl import load_workbook

    b_ids = claim_ids(client, tokens[T2_CSUITE])
    r = client.get("/api/v1/export/claims.xlsx", headers=auth(tokens[P1_CSUITE]))
    assert r.status_code == 200

    wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    text = "\n".join(
        "" if v is None else str(v)
        for ws in wb.worksheets for row in ws.iter_rows(values_only=True) for v in row
    )
    wb.close()
    leaked = [cid for cid in b_ids if cid in text]
    assert not leaked, f"export leaked {len(leaked)} claims from the other tenant"


# ── FNOL: a site in another tenant cannot be used ──────────────────────────

def test_cannot_raise_fnol_against_another_tenants_site(client, tokens):
    r = client.post(
        "/api/v1/fnol",
        json={"site_org_node": "SITE-NW-LEEDS", "date_of_loss": "2026-05-01",
              "claim_type": "Claim", "product_line": "Property & Equipment",
              "loss_description": "cross-tenant attempt", "values": {}},
        headers={**auth(tokens[P5_BISTRO_MGR]), "Idempotency-Key": "xt-fnol-1"},
    )
    assert r.status_code == 403


def test_policies_are_not_offered_from_another_tenant(client, tokens):
    r = client.get("/api/v1/policies", headers=auth(tokens[T2_STORE_MGR]))
    assert r.status_code == 200
    for p in r.json().get("items", r.json() if isinstance(r.json(), list) else []):
        node = p.get("org_node", "")
        assert "JFK" not in node and "LHR" not in node and "SIN" not in node


# ── the second barrier, tested on its own ─────────────────────────────────
#
# Everything above passes even with the downstream tenant predicate removed, because
# resolve_authorized_scope already filters descendants by client_id - so sp.scope only
# ever holds same-tenant nodes and the node predicate alone is sufficient.
#
# That is exactly what defence in depth means, and exactly why it needs its own tests.
# The cases below corrupt the first barrier deliberately and assert that the second one
# still holds. Without these, "two independent predicates" would be an untested claim.

def test_misstamped_claim_is_excluded_by_the_tenant_predicate(client, tokens):
    """
    A claim sitting on tenant A's node but stamped to tenant B must not be returned.

    This is the realistic corruption: an MDM re-parent or a bad import leaves the row's
    org_node and client_id disagreeing. Node scope alone would surface it.
    """
    from app.db import execute, query_one

    node = "SITE-JFK-T4-BISTRO"          # belongs to CLIENT_A
    execute(
        """INSERT INTO claims (aon_claim_id, org_node, client_id, status, claim_type,
                               is_draft, global_product, date_of_loss, gross_incurred,
                               total_paid, total_outstanding, currency_code,
                               restricted_access)
           VALUES ('XT-MISSTAMP', :n, :c, 'Open', 'Claim', 0, 'Property & Equipment',
                   '2026-05-01', 1234.0, 0, 1234.0, 'USD', 0)""",
        {"n": node, "c": CLIENT_B},
    )
    try:
        visible = claim_ids(client, tokens[P1_CSUITE])
        assert "XT-MISSTAMP" not in visible, (
            "a claim stamped to another tenant was returned because it sat on an "
            "in-scope node - the tenant predicate is not being applied"
        )
        assert client.get("/api/v1/claims/XT-MISSTAMP",
                          headers=auth(tokens[P1_CSUITE])).status_code == 403
    finally:
        execute("DELETE FROM claims WHERE aon_claim_id = 'XT-MISSTAMP'")
        assert query_one(
            "SELECT COUNT(*) AS n FROM claims WHERE aon_claim_id = 'XT-MISSTAMP'"
        )["n"] == 0


def test_misstamped_claim_is_excluded_from_export(client, tokens):
    """The same corruption must not reach a generated file either."""
    from openpyxl import load_workbook

    from app.db import execute

    execute(
        """INSERT INTO claims (aon_claim_id, org_node, client_id, status, claim_type,
                               is_draft, global_product, date_of_loss, gross_incurred,
                               total_paid, total_outstanding, currency_code,
                               restricted_access)
           VALUES ('XT-MISSTAMP-2', 'SITE-JFK-T4-BISTRO', :c, 'Open', 'Claim', 0,
                   'Cyber', '2026-05-02', 999.0, 0, 999.0, 'USD', 0)""",
        {"c": CLIENT_B},
    )
    try:
        r = client.get("/api/v1/export/claims.xlsx", headers=auth(tokens[P1_CSUITE]))
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        text = "\n".join(
            "" if v is None else str(v)
            for ws in wb.worksheets for row in ws.iter_rows(values_only=True) for v in row
        )
        wb.close()
        assert "XT-MISSTAMP-2" not in text
    finally:
        execute("DELETE FROM claims WHERE aon_claim_id = 'XT-MISSTAMP-2'")


def test_misstamped_saved_view_is_excluded(client, tokens):
    """A shared view stamped to another tenant must not surface on node scope alone."""
    from app.db import execute

    execute(
        """INSERT INTO saved_views (view_id, owner_sub, owner_name, org_node, client_id,
                                    name, filters_json, is_shared, created_at)
           VALUES ('xt-view', 'poc|other', 'Other', 'LOC-JFK', :c,
                   'Mis-stamped', '{}', 1, '2026-05-01T00:00:00')""",
        {"c": CLIENT_B},
    )
    try:
        seen = {v["view_id"] for v in
                client.get("/api/v1/views", headers=auth(tokens[P1_CSUITE])).json()["items"]}
        assert "xt-view" not in seen
    finally:
        execute("DELETE FROM saved_views WHERE view_id = 'xt-view'")


# ── role held constant, tenant varied ─────────────────────────────────────

def test_same_role_different_tenant_sees_different_data(client, tokens):
    """
    P3 and T2_REGIONAL are both location managers with identical privileges.

    Holding the role constant means a difference in what they see can only be the
    tenant boundary, not an entitlement.
    """
    a = claim_ids(client, tokens[P3_JFK_DIRECTOR])
    b = claim_ids(client, tokens[T2_REGIONAL])
    assert a and b
    assert not (a & b)
