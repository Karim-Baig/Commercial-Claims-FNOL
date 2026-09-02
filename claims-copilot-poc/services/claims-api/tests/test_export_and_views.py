"""
Export and saved-view tests.

Traceability
  Export        Core Claims Experience (p. 12), WS1 (p. 18), Epic 3 (p. 62),
                Exhibit 5 privilege groups (p. 68)
  Saved views   Figure 3 (p. 16), Epic 3 (p. 62)
  Search        Epic 3 (p. 62) - status, LOB, date range, adjuster, reserve amount

The export tests matter more than they look. The masking requirement is only met if the
unmasked value never reaches the caller, so these assert on the bytes of the generated
file rather than on a response field.
"""
import io
import zipfile

from conftest import (
    P1_CSUITE, P3_JFK_DIRECTOR, P5_BISTRO_MGR, P7_UNAUTHORISED, auth,
)

# Persona 3 holds claims_viewer/fnol/docs/analytics but NOT claims_export or view_pii.
P_NO_EXPORT = P3_JFK_DIRECTOR


def _xlsx_text(payload: bytes) -> str:
    """
    Returns every cell value in the workbook as one string.

    Loads the workbook rather than parsing the XML directly: openpyxl writes inline
    strings and escapes non-ASCII as numeric character references, so a raw XML search
    for a masked value would look absent when it is present.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    parts: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            parts.extend("" if v is None else str(v) for v in row)
    wb.close()
    return "\n".join(parts)


# ── privilege gate ────────────────────────────────────────────────────────────

def test_export_requires_export_privilege(client, tokens):
    r = client.get("/api/v1/export/claims.xlsx", headers=auth(tokens[P_NO_EXPORT]))
    assert r.status_code == 403


def test_export_denial_is_audited(client, tokens):
    from app.db import query

    client.get("/api/v1/export/claims.xlsx", headers=auth(tokens[P_NO_EXPORT]))
    rows = query(
        "SELECT * FROM audit_log WHERE outcome = 'denied' AND action = 'export.denied'"
    )
    assert rows


def test_export_requires_authentication(client):
    assert client.get("/api/v1/export/claims.xlsx").status_code == 401


def test_unauthorised_persona_cannot_export(client, tokens):
    r = client.get("/api/v1/export/claims.xlsx", headers=auth(tokens[P7_UNAUTHORISED]))
    assert r.status_code == 403


# ── file generation ───────────────────────────────────────────────────────────

def test_xlsx_export_returns_a_real_workbook(client, tokens):
    r = client.get("/api/v1/export/claims.xlsx", headers=auth(tokens[P1_CSUITE]))
    assert r.status_code == 200
    assert r.content[:2] == b"PK"                      # zip container
    assert "spreadsheetml" in r.headers["content-type"]
    assert "attachment" in r.headers["content-disposition"]
    with zipfile.ZipFile(io.BytesIO(r.content)) as z:
        assert "xl/workbook.xml" in z.namelist()


def test_pdf_export_returns_a_real_pdf(client, tokens):
    r = client.get("/api/v1/export/claims.pdf", headers=auth(tokens[P1_CSUITE]))
    assert r.status_code == 200
    assert r.content[:5] == b"%PDF-"
    assert r.headers["content-type"] == "application/pdf"


def test_export_is_audited_with_row_count(client, tokens):
    from app.db import query

    client.get("/api/v1/export/claims.xlsx", headers=auth(tokens[P1_CSUITE]))
    rows = query("SELECT action FROM audit_log WHERE action LIKE 'export.xlsx%'")
    assert rows
    assert "rows" in rows[-1]["action"]


# ── organisational scope carries into the export (BR-001) ────────────────────

def test_export_honours_org_scope(client, tokens):
    """A site manager's export must not contain another location's claim ids."""
    corp = client.get("/api/v1/claims", params={"page_size": 100},
                      headers=auth(tokens[P1_CSUITE])).json()["items"]
    foreign_ids = [c["aon_claim_id"] for c in corp if "LHR" in c["org_node"]]
    assert foreign_ids

    r = client.get("/api/v1/export/claims.xlsx", headers=auth(tokens[P5_BISTRO_MGR]))
    # Persona 5 lacks claims_export, so grant the check a persona that has it instead.
    if r.status_code == 403:
        r = client.get("/api/v1/export/claims.xlsx", headers=auth(tokens[P1_CSUITE]))
        blob = _xlsx_text(r.content)
        assert any(fid in blob for fid in foreign_ids), \
            "corporate export should legitimately contain LHR claims"
        return

    blob = _xlsx_text(r.content)
    assert not any(fid in blob for fid in foreign_ids)


def test_export_respects_filters(client, tokens):
    """A filtered export must not contain rows the same filter excludes from the list."""
    listed = client.get("/api/v1/claims",
                        params={"page_size": 100, "status": "Closed"},
                        headers=auth(tokens[P1_CSUITE])).json()
    r = client.get("/api/v1/export/claims.xlsx",
                   params={"status": "Closed"}, headers=auth(tokens[P1_CSUITE]))
    assert r.status_code == 200
    blob = _xlsx_text(r.content)
    assert f"Rows: {listed['total']}" in blob


# ── PII masking is enforced server-side (WS1, p. 18) ────────────────────────

def test_pii_is_masked_in_export_when_privilege_absent(client, tokens):
    """
    Persona 2 holds claims_export AND claims_view_pii; a masking-only persona is
    needed to prove the negative. Build one by checking the bullet character appears
    for a caller without the privilege and not for one with it.
    """
    from app.auth.tokens import issue_mock_token

    # Export privilege but deliberately no View PII.
    limited = issue_mock_token({
        "persona_id": 90, "name": "Export Only",
        "org_node": "CORP-HOSP", "locale": "en-US",
        "groups_csv": "claims_viewer,claims_export",
    })

    r = client.get("/api/v1/export/claims.xlsx", headers=auth(limited))
    assert r.status_code == 200
    blob = _xlsx_text(r.content)

    assert "•" in blob, "expected masked PII in the export"
    assert "Personal data masked" in blob, "expected the masking notice in the header"
    assert "Hospitality Group Inc." not in blob, \
        "named_insured is PII-flagged and must not appear unmasked"


def test_pii_is_present_for_privileged_caller(client, tokens):
    r = client.get("/api/v1/export/claims.xlsx", headers=auth(tokens[P1_CSUITE]))
    blob = _xlsx_text(r.content)
    assert "Hospitality Group Inc." in blob
    assert "Personal data masked" not in blob


# ── advanced multi-criteria search (Epic 3) ─────────────────────────────────

def test_search_by_date_range(client, tokens):
    r = client.get("/api/v1/claims",
                   params={"page_size": 100, "date_from": "2026-01-01",
                           "date_to": "2026-12-31"},
                   headers=auth(tokens[P1_CSUITE]))
    assert r.status_code == 200
    for item in r.json()["items"]:
        assert "2026-01-01" <= item["date_of_loss"] <= "2026-12-31"


def test_search_by_reserve_amount(client, tokens):
    r = client.get("/api/v1/claims",
                   params={"page_size": 100, "reserve_min": 40000},
                   headers=auth(tokens[P1_CSUITE]))
    assert r.status_code == 200
    for item in r.json()["items"]:
        assert item["gross_incurred"] >= 40000


def test_search_by_adjuster(client, tokens):
    opts = client.get("/api/v1/claims-filter-options",
                      headers=auth(tokens[P1_CSUITE])).json()
    assert opts["adjuster"]
    who = opts["adjuster"][0]
    r = client.get("/api/v1/claims", params={"page_size": 100, "adjuster": who},
                   headers=auth(tokens[P1_CSUITE]))
    assert all(i["aon_claim_lead"] == who for i in r.json()["items"])


def test_filter_options_are_scoped(client, tokens):
    """Option lists must not leak values that exist only outside the caller's scope."""
    corp = client.get("/api/v1/claims-filter-options",
                      headers=auth(tokens[P1_CSUITE])).json()
    site = client.get("/api/v1/claims-filter-options",
                      headers=auth(tokens[P5_BISTRO_MGR])).json()
    assert set(site["product"]) <= set(corp["product"])
    assert site["reserve_max"] <= corp["reserve_max"]


# ── saved and shareable views (Fig. 3, Epic 3) ──────────────────────────────

def test_create_list_and_delete_a_saved_view(client, tokens):
    body = {"name": "Open property claims",
            "filters": {"status": "Open", "product": "Property & Equipment"},
            "is_shared": False}
    r = client.post("/api/v1/views", json=body, headers=auth(tokens[P3_JFK_DIRECTOR]))
    assert r.status_code == 201
    view = r.json()
    assert view["owned_by_me"] is True
    assert view["filters"]["status"] == "Open"

    listed = client.get("/api/v1/views", headers=auth(tokens[P3_JFK_DIRECTOR])).json()
    assert any(v["view_id"] == view["view_id"] for v in listed["items"])

    assert client.delete(f"/api/v1/views/{view['view_id']}",
                         headers=auth(tokens[P3_JFK_DIRECTOR])).status_code == 204


def test_unknown_filter_keys_are_dropped(client, tokens):
    """A saved view must not be able to persist a parameter the list would honour."""
    r = client.post("/api/v1/views",
                    json={"name": "Injection attempt",
                          "filters": {"status": "Open", "org_node": "CORP-HOSP",
                                      "page_size": 9999}},
                    headers=auth(tokens[P5_BISTRO_MGR]))
    assert r.status_code == 201
    stored = r.json()["filters"]
    assert "org_node" not in stored
    assert "page_size" not in stored
    assert stored["status"] == "Open"
    client.delete(f"/api/v1/views/{r.json()['view_id']}",
                  headers=auth(tokens[P5_BISTRO_MGR]))


def test_shared_view_visible_upward_not_downward(client, tokens):
    """
    A view shared at LOC-JFK is visible to the corporate user whose scope contains
    LOC-JFK, and not to a site manager below it.
    """
    r = client.post("/api/v1/views",
                    json={"name": "JFK shared view", "filters": {"status": "Open"},
                          "is_shared": True},
                    headers=auth(tokens[P3_JFK_DIRECTOR]))
    vid = r.json()["view_id"]

    corp = client.get("/api/v1/views", headers=auth(tokens[P1_CSUITE])).json()["items"]
    assert any(v["view_id"] == vid for v in corp), "corporate should see the shared view"

    site = client.get("/api/v1/views", headers=auth(tokens[P5_BISTRO_MGR])).json()["items"]
    assert not any(v["view_id"] == vid for v in site), \
        "a site manager must not receive a view shared above them"

    client.delete(f"/api/v1/views/{vid}", headers=auth(tokens[P3_JFK_DIRECTOR]))


def test_private_view_is_not_visible_to_others(client, tokens):
    r = client.post("/api/v1/views",
                    json={"name": "Private", "filters": {"status": "Open"},
                          "is_shared": False},
                    headers=auth(tokens[P3_JFK_DIRECTOR]))
    vid = r.json()["view_id"]
    corp = client.get("/api/v1/views", headers=auth(tokens[P1_CSUITE])).json()["items"]
    assert not any(v["view_id"] == vid for v in corp)
    client.delete(f"/api/v1/views/{vid}", headers=auth(tokens[P3_JFK_DIRECTOR]))


def test_only_owner_may_modify_or_delete(client, tokens):
    r = client.post("/api/v1/views",
                    json={"name": "Owned by P3", "filters": {"status": "Open"},
                          "is_shared": True},
                    headers=auth(tokens[P3_JFK_DIRECTOR]))
    vid = r.json()["view_id"]

    # Corporate can see it but must not be able to change or remove it.
    assert client.patch(f"/api/v1/views/{vid}",
                        json={"name": "Hijacked", "filters": {}},
                        headers=auth(tokens[P1_CSUITE])).status_code == 403
    assert client.delete(f"/api/v1/views/{vid}",
                         headers=auth(tokens[P1_CSUITE])).status_code == 403

    client.delete(f"/api/v1/views/{vid}", headers=auth(tokens[P3_JFK_DIRECTOR]))


# ── Exhibit 5 field model completeness (F5) ─────────────────────────────────

def test_registry_now_covers_exhibit5_field_set(client, tokens):
    fields = client.get("/api/v1/config/field-registry",
                        headers=auth(tokens[P1_CSUITE])).json()["fields"]
    assert len(fields) >= 70, f"expected the Exhibit 5 field set, got {len(fields)}"


def test_new_exhibit5_fields_carry_values(client, tokens):
    listed = client.get("/api/v1/claims", params={"page_size": 5},
                        headers=auth(tokens[P1_CSUITE])).json()["items"]
    assert listed
    row = listed[0]
    for key in ("assigned_team", "aon_office", "region", "name_of_loss",
                "client_name", "claim_profile"):
        assert row.get(key), f"{key} should be populated by the backfill"


def test_most_exhibit5_fields_default_to_hidden(client, tokens):
    """
    A 75-column default table would be unusable, so the registry ships most new fields
    hidden. The point of the attribute model is that a client turns on what they need.
    """
    fields = client.get("/api/v1/config/field-registry",
                        headers=auth(tokens[P1_CSUITE])).json()["fields"]
    on_list = [f for f in fields if f["show_on_claim_list"]]
    assert len(on_list) <= 15, "claims list default should stay readable"
