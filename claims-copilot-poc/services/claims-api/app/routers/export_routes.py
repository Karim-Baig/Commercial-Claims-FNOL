"""
Claims list and claim record export.

Requirement: "Bulk export and print of the claims list or claim detail to an Aon
branded client deliverable in Excel or PDF" - Functional Requirements, Core Claims
Experience (p. 12); Workstream 1 (p. 18) adds "respecting field-level entitlement and
PII masking"; Epic 3 (p. 62) lists Bulk Export & Print.

Why this is generated server-side
---------------------------------
The masking requirement is the reason. If the browser received unmasked rows and
redacted them for display, the data would already have left the trust boundary - the
masking would be cosmetic and a network trace would defeat it. Generating the file here
means a caller without the View PII privilege never receives the underlying values in
any form.

Four gates apply, in this order:

  1. claims_export privilege                     (Exhibit 5 privilege groups, p. 68)
  2. organisational scope                        (BR-001, p. 38)
  3. registry field selection                    (Exhibit 5 attribute model, p. 68)
  4. PII masking per field where privilege absent (WS1, p. 18)

Every export is written to the audit log with the row count, because a bulk extract of
Confidential data is exactly the event an auditor will ask about.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, Response

from ..auth.scope import ScopedPrincipal, current_scope
from ..db import query
from ..services import audit
from .claims_routes import CLAIM_SELECT, ClaimFilters, build_claim_where

router = APIRouter(prefix="/export", tags=["export"])

# NFR-34 payload protection. A bulk extract is still an interface that can be misused.
MAX_EXPORT_ROWS = 5000

AON_NAVY = "0F2B5B"
AON_GREY = "F1F4F7"


# ─────────────────────────────────────────────────── shared preparation
def _resolve_columns(surface: str) -> list[dict]:
    """
    Field selection from the registry (Exhibit 5), not a hard-coded list.

    Uses the same attribute model that drives the on-screen table, so an exported file
    contains exactly the columns the user was looking at - including a field an
    administrator has just switched on (NFR-45).
    """
    col = "show_on_claim_record" if surface == "record" else "show_on_claim_list"
    return query(
        f"""SELECT field_key, label_token, is_pii, value_type, c2s_order
            FROM field_registry
            WHERE available_in_meridian = 1 AND {col} = 1
            ORDER BY c2s_order"""
    )


def _mask(value) -> str:
    """
    Redacts a PII value while leaving it recognisable enough to be useful.

    Mirrors the client-side maskPii helper so an export and the screen agree.
    """
    s = "" if value is None else str(value)
    if not s:
        return ""
    return " ".join(
        w if len(w) <= 1 else w[0] + "•" * min(len(w) - 1, 6) for w in s.split(" ")
    )


def _cell(row: dict, field: dict, can_view_pii: bool):
    raw = row.get(field["field_key"])
    if field["is_pii"] and not can_view_pii:
        return _mask(raw)
    if raw is None:
        return ""
    if field["value_type"] in ("money", "number"):
        return raw          # keep numeric so Excel can total the column
    if field["field_key"] in ("escalated", "disputed_claim"):
        return "Yes" if raw else "No"
    return str(raw)


def _fetch(sp: ScopedPrincipal, filters: ClaimFilters, sort: str, dir_: str):
    where_sql, params = build_claim_where(sp, filters)
    params["lim"] = MAX_EXPORT_ROWS
    order = sort if sort else "date_of_loss"
    return query(
        f"""SELECT {CLAIM_SELECT} FROM claims WHERE {where_sql}
            ORDER BY {order} {'ASC' if dir_ == 'asc' else 'DESC'}
            LIMIT :lim""",
        params,
    )


def _label(field: dict) -> str:
    """
    Column caption.

    Export runs outside a React tree so it cannot call the UI translator. The token is
    de-slugged into a readable heading; wiring the server to the same resource bundles
    is the correct next step for a localised export.
    """
    token = field["label_token"].replace("field.", "").replace("_", " ")
    return token.title()


def _guard(sp: ScopedPrincipal) -> None:
    if not sp.has("claims_export"):
        audit.log_denied(sp.sub, "export.denied", "claims", "list")
        raise HTTPException(403, "Export Data entitlement not held")


def _filters_from_query(**kw) -> ClaimFilters:
    return ClaimFilters(**{k: v for k, v in kw.items() if v is not None})


def _stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")


# ─────────────────────────────────────────────────── Excel
@router.get("/claims.xlsx")
def export_claims_xlsx(
    sp: ScopedPrincipal = Depends(current_scope),
    tab: str = Query("submitted", pattern="^(submitted|drafts)$"),
    q: str | None = None,
    status: str | None = None,
    sub_status: str | None = None,
    product: str | None = None,
    product_category: str | None = None,
    adjuster: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    reserve_min: float | None = None,
    reserve_max: float | None = None,
    claim_type: str | None = None,
    sort: str = "date_of_loss",
    dir: str = Query("desc", pattern="^(asc|desc)$"),
):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _guard(sp)
    filters = _filters_from_query(
        tab=tab, q=q, status=status, sub_status=sub_status, product=product,
        product_category=product_category, adjuster=adjuster, date_from=date_from,
        date_to=date_to, reserve_min=reserve_min, reserve_max=reserve_max,
        claim_type=claim_type,
    )
    fields = _resolve_columns("list")
    rows = _fetch(sp, filters, sort, dir)
    can_pii = sp.has("claims_view_pii")

    wb = Workbook()
    ws = wb.active
    ws.title = "Claims"

    # Aon-branded header band (p. 12 requires an Aon branded client deliverable)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(fields), 4))
    title = ws.cell(row=1, column=1, value="Aon  |  Claims Export")
    title.font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=AON_NAVY)
    title.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(fields), 4))
    meta = (
        f"Scope: {sp.principal.org_node}   ·   "
        f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}   ·   "
        f"Rows: {len(rows)}"
    )
    if not can_pii:
        meta += "   ·   Personal data masked (View PII entitlement not held)"
    m = ws.cell(row=2, column=1, value=meta)
    m.font = Font(name="Calibri", size=9, italic=True, color="3E4C59")
    m.fill = PatternFill("solid", fgColor=AON_GREY)

    for c, f in enumerate(fields, start=1):
        cell = ws.cell(row=4, column=c, value=_label(f))
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=AON_NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = (
            34 if f["value_type"] == "text" and "description" in f["field_key"] else 18
        )
    ws.freeze_panes = "A5"

    for r, row in enumerate(rows, start=5):
        for c, f in enumerate(fields, start=1):
            cell = ws.cell(row=r, column=c, value=_cell(row, f, can_pii))
            cell.font = Font(name="Calibri", size=10)
            if f["value_type"] == "money":
                cell.number_format = "#,##0.00"

    ws.auto_filter.ref = (
        f"A4:{get_column_letter(len(fields))}{max(4 + len(rows), 5)}"
    )

    buf = io.BytesIO()
    wb.save(buf)
    audit.log(sp.sub, f"export.xlsx({len(rows)} rows)", "claims", "list",
              sp.principal.org_node)

    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
                f'attachment; filename="aon-claims-{_stamp()}.xlsx"'
        },
    )


# ─────────────────────────────────────────────────── PDF
@router.get("/claims.pdf")
def export_claims_pdf(
    sp: ScopedPrincipal = Depends(current_scope),
    tab: str = Query("submitted", pattern="^(submitted|drafts)$"),
    q: str | None = None,
    status: str | None = None,
    sub_status: str | None = None,
    product: str | None = None,
    product_category: str | None = None,
    adjuster: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    reserve_min: float | None = None,
    reserve_max: float | None = None,
    claim_type: str | None = None,
    sort: str = "date_of_loss",
    dir: str = Query("desc", pattern="^(asc|desc)$"),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    _guard(sp)
    filters = _filters_from_query(
        tab=tab, q=q, status=status, sub_status=sub_status, product=product,
        product_category=product_category, adjuster=adjuster, date_from=date_from,
        date_to=date_to, reserve_min=reserve_min, reserve_max=reserve_max,
        claim_type=claim_type,
    )
    # A landscape page cannot carry 25 columns legibly, so the PDF takes the first
    # eight registry columns. Excel remains the full-fidelity export.
    fields = _resolve_columns("list")[:8]
    rows = _fetch(sp, filters, sort, dir)
    can_pii = sp.has("claims_view_pii")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title="Aon Claims Export", author="Aon",
    )

    navy = colors.HexColor("#0F2B5B")
    grey = colors.HexColor("#F1F4F7")

    h = ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=15,
                       textColor=navy, spaceAfter=2)
    sub = ParagraphStyle("s", fontName="Helvetica-Oblique", fontSize=8,
                         textColor=colors.HexColor("#3E4C59"), spaceAfter=8)

    meta = (
        f"Scope: {sp.principal.org_node} &nbsp;&middot;&nbsp; "
        f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')} "
        f"&nbsp;&middot;&nbsp; Rows: {len(rows)}"
    )
    if not can_pii:
        meta += " &nbsp;&middot;&nbsp; Personal data masked (View PII entitlement not held)"

    cellstyle = ParagraphStyle("c", fontName="Helvetica", fontSize=6.5, leading=8)
    headstyle = ParagraphStyle("hd", fontName="Helvetica-Bold", fontSize=6.5,
                               leading=8, textColor=colors.white)

    data = [[Paragraph(_label(f), headstyle) for f in fields]]
    for row in rows:
        data.append([
            Paragraph(str(_cell(row, f, can_pii))[:90], cellstyle) for f in fields
        ])

    table = Table(data, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD2D9")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, grey]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))

    story = [Paragraph("Aon &nbsp;|&nbsp; Claims Export", h), Paragraph(meta, sub)]
    if rows:
        story.append(table)
    else:
        story.append(Paragraph("No claims match the current filters.", cellstyle))
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(
        "Confidential. Contains Aon client claims data. Distribute only in line with "
        "your organisation's data handling policy.", sub))

    doc.build(story)
    audit.log(sp.sub, f"export.pdf({len(rows)} rows)", "claims", "list",
              sp.principal.org_node)

    return Response(
        buf.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                f'attachment; filename="aon-claims-{_stamp()}.pdf"'
        },
    )
